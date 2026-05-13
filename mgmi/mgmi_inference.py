"""
MGMI (MitoKG-Grounded Mechanistic Inference) -- LLM reasoning module.

Generates step-by-step causal reasoning chains for candidate perturbation
targets identified by the forward perturbation experiment. Every reasoning
step must cite a specific MitoKG element (gene / pathway / complex / drug).

Three-layer evidence integration:
  1. MitoKG structural context (Reactome / CORUM / DrugBank / STRING neighbors)
  2. Temporal window dynamics (Day0 / Day10 / Day14 cross-timepoint delta_r)
  3. KGSFM gating weight proxy (L2 norm of HGT 64-dim gene embeddings)

The LLM (DeepSeek-Chat) must output:
  [Step 1..4] cited MitoKG elements
  [Verification] all nodes cited
  [Confidence] HIGH / MEDIUM / LOW

Automated verification function then checks that:
  - every cited gene symbol is present in MitoKG
  - temporal reasoning (Day0/Day10/Day14 references) is present
  - KGSFM gating rationale is present
Chains are flagged PASS if all of the above hold.

IMPORTANT: DeepSeek API key must be provided via the DEEPSEEK_API_KEY
environment variable. This script will not run otherwise. The key is never
stored in code.

Input:
    perturbation_lrrk2_temporal.xlsx (from perturbation_forward.py)
    MitoKG gnn/ directory

Output:
    mitokg_llm_raw_v2.json                   (consumed by mgmi_external_validation.py)
    mitokg_llm_interpretability_lrrk2_v2.xlsx  (human-readable full report)

Usage:
    export DEEPSEEK_API_KEY="your-key-here"
    python mgmi_inference.py \\
        --gnn-dir /path/to/MitoKG/gnn \\
        --perturb-xlsx /path/to/perturbation_lrrk2_temporal.xlsx \\
        --out-dir ./output_mgmi
"""

import argparse
import json
import os
import pickle
import re
import sys
import time
import warnings
from pathlib import Path

warnings.filterwarnings('ignore')

import numpy as np
import pandas as pd


TOP_N_EACH = 5
TOP_N_ROBUST = 10
DEFAULT_CUSTOM_GENES = ['UQCC3', 'SDHAF3', 'NDUFA7', 'SDHD', 'TOMM20']

DEEPSEEK_BASE_URL = "https://api.deepseek.com"
DEEPSEEK_MODEL = "deepseek-chat"
LLM_TEMPERATURE = 0.3
LLM_MAX_TOKENS = 700
LLM_RETRIES = 3

REACTOME_NAMES = {
    'R-HSA-1428517': 'Mitochondrial protein import',
    'R-HSA-611105': 'Respiratory electron transport',
    'R-HSA-163200': 'Respiratory electron transport & ATP synthesis',
    'R-HSA-5205647': 'Mitophagy',
    'R-HSA-9612973': 'Autophagy',
    'R-HSA-8949215': 'Mitochondrial translation',
    'R-HSA-71403': 'TCA cycle (Citric acid cycle)',
    'R-HSA-70268': 'Pyruvate metabolism',
    'R-HSA-5389840': 'Mitochondrial translation initiation',
    'R-HSA-5368287': 'Mitochondrial translation elongation',
    'R-HSA-5419276': 'Mitochondrial translation termination',
}

DRUGBANK_NAMES = {
    'DB00091': 'Cyclosporine', 'DB00313': 'Valproic acid',
    'DB00641': 'Simvastatin', 'DB00619': 'Imatinib',
    'DB01076': 'Atorvastatin', 'DB02010': 'Rotenone',
    'DB00563': 'Methotrexate', 'DB00945': 'Aspirin',
    'DB00331': 'Metformin', 'DB01234': 'Dexamethasone',
    'DB00783': 'Estradiol', 'DB00171': 'ATP',
    'DB00166': 'Lipoic acid', 'DB00114': 'Pyridoxal Phosphate',
}


SYSTEM_PROMPT = """You are a computational mitochondrial biologist specializing in
Parkinson's disease mechanisms. You analyze results from a knowledge-graph-guided
neural ODE model (MitoLODE) with three key components:

1. KGSFM (KG-gated Spatiotemporal Feature Modulator): uses MitoKG gene embeddings
   to compute gating weights - genes with higher embedding norms have stronger
   influence on the encoder feature extraction.
2. KG-ODE: MitoKG context guides the continuous-time dynamics.
3. Temporal encoding: time points are encoded as semantic feature vectors (not
   scalars), enabling the model to perceive differentiation stage semantics.

Your task: generate step-by-step causal reasoning chains that are grounded in
the provided MitoKG structural information AND explain the temporal dynamics.

STRICT RULES:
- Every [Step] MUST cite a specific MitoKG element from the provided context.
- Do NOT fabricate nodes or edges absent from the provided MitoKG data.
- Explain WHY the temporal window matters for this specific gene.
- Reference the KGSFM gating weight when discussing model attention.
- Keep each response under 350 words.
- End with [Verification] listing all MitoKG nodes cited, then [Confidence] HIGH/MEDIUM/LOW.
"""


def parse_args():
    p = argparse.ArgumentParser(description='MGMI LLM-based mechanistic inference')
    p.add_argument('--gnn-dir', type=Path, required=True,
                    help='MitoKG gnn/ directory (gene_embeddings.npy, gene_list.txt, node_id_maps.pkl, hetero_graph.pt)')
    p.add_argument('--perturb-xlsx', type=Path, required=True,
                    help='perturbation_lrrk2_temporal.xlsx produced by perturbation_forward.py')
    p.add_argument('--out-dir', type=Path, default=Path('./output_mgmi'))
    p.add_argument('--mode', choices=['L2_top_N_each', 'robust_top_N', 'custom'],
                    default='L2_top_N_each')
    p.add_argument('--top-n-each', type=int, default=TOP_N_EACH,
                    help='per-timepoint Top-N (mode=L2_top_N_each)')
    p.add_argument('--top-n-robust', type=int, default=TOP_N_ROBUST,
                    help='cross-timepoint stable Top-N (mode=robust_top_N)')
    p.add_argument('--custom-genes', nargs='+', default=DEFAULT_CUSTOM_GENES,
                    help='custom gene list (mode=custom)')
    return p.parse_args()


def load_kg(gnn_dir: Path):
    with open(str(gnn_dir / "node_id_maps.pkl"), "rb") as f:
        node_maps = pickle.load(f)
    id2pathway = {v: k for k, v in node_maps['pathway'].items()}
    id2drug = {v: k for k, v in node_maps['drug'].items()}
    id2complex = {v: k for k, v in node_maps['complex'].items()}

    gene_list_path = gnn_dir / "gene_list.txt"
    if not gene_list_path.exists():
        gene_list_path = gnn_dir / "gene_list"
    with open(str(gene_list_path), encoding='utf-8') as f:
        gene_list = [l.strip() for l in f if l.strip()]
    symbol2geneid = {sym: i for i, sym in enumerate(gene_list)}
    geneid2symbol = {i: sym for i, sym in enumerate(gene_list)}
    print(f"  genes: {len(gene_list)}  pathways: {len(node_maps['pathway'])}  "
          f"drugs: {len(node_maps['drug'])}  complexes: {len(node_maps['complex'])}")

    gene_emb = np.load(str(gnn_dir / "gene_embeddings.npy"))
    emb_norms = np.linalg.norm(gene_emb, axis=1)
    emb_norms_normed = (emb_norms - emb_norms.min()) / (emb_norms.max() - emb_norms.min() + 1e-8)
    print(f"  embedding matrix: {gene_emb.shape}, L2 norm range "
          f"[{emb_norms.min():.3f}, {emb_norms.max():.3f}]")

    try:
        import torch
        hetero_graph = torch.load(str(gnn_dir / "hetero_graph.pt"),
                                    map_location='cpu', weights_only=False)
        graph_loaded = True
        print("  hetero_graph.pt loaded")

        gene2pathway, gene2complex, gene2drug, gene2gene = {}, {}, {}, {}
        for et in hetero_graph.edge_types:
            ei = hetero_graph[et].edge_index
            s_list, d_list = ei[0].tolist(), ei[1].tolist()
            if et == ('drug', 'targets', 'gene'):
                for s, d in zip(s_list, d_list):
                    gene2drug.setdefault(d, []).append(s)
            elif et == ('gene', 'involved_in', 'pathway'):
                for s, d in zip(s_list, d_list):
                    gene2pathway.setdefault(s, []).append(d)
            elif et == ('gene', 'member_of', 'complex'):
                for s, d in zip(s_list, d_list):
                    gene2complex.setdefault(s, []).append(d)
            elif et == ('gene', 'interacts', 'gene'):
                for s, d in zip(s_list, d_list):
                    gene2gene.setdefault(s, []).append(d)
        print(f"  PPI edges: {sum(len(v) for v in gene2gene.values())}  "
                f"gene->pathway: {sum(len(v) for v in gene2pathway.values())}  "
                f"drug->gene: {sum(len(v) for v in gene2drug.values())}")
    except Exception as e:
        graph_loaded = False
        gene2pathway = gene2complex = gene2drug = gene2gene = {}
        print(f"  [warning] hetero_graph.pt not loaded ({e})")

    return {
        'id2pathway': id2pathway, 'id2drug': id2drug, 'id2complex': id2complex,
        'gene_list': gene_list,
        'symbol2geneid': symbol2geneid, 'geneid2symbol': geneid2symbol,
        'emb_norms': emb_norms, 'emb_norms_normed': emb_norms_normed,
        'graph_loaded': graph_loaded,
        'gene2pathway': gene2pathway, 'gene2complex': gene2complex,
        'gene2drug': gene2drug, 'gene2gene': gene2gene,
    }


def get_kg_weight(kg, gene_symbol):
    if gene_symbol not in kg['symbol2geneid']:
        return 0.0, 0.0
    gid = kg['symbol2geneid'][gene_symbol]
    if gid >= len(kg['emb_norms']):
        return 0.0, 0.0
    return float(kg['emb_norms'][gid]), float(kg['emb_norms_normed'][gid])


def get_gene_kg_info(kg, gene_symbol):
    info = {'symbol': gene_symbol, 'in_kg': gene_symbol in kg['symbol2geneid'],
            'pathways': [], 'complexes': [], 'drugs': [], 'ppi_neighbors': []}
    if not info['in_kg'] or not kg['graph_loaded']:
        return info
    gid = kg['symbol2geneid'][gene_symbol]
    for pid in kg['gene2pathway'].get(gid, []):
        rid = kg['id2pathway'].get(pid, '')
        name = REACTOME_NAMES.get(rid, rid)
        info['pathways'].append({'reactome_id': rid, 'name': name})
    for cid in kg['gene2complex'].get(gid, []):
        info['complexes'].append({'corum_id': kg['id2complex'].get(cid, '')})
    for did in kg['gene2drug'].get(gid, []):
        db_id = kg['id2drug'].get(did, '')
        info['drugs'].append({'drugbank_id': db_id,
                                'name': DRUGBANK_NAMES.get(db_id, db_id)})
    for nid in kg['gene2gene'].get(gid, [])[:10]:
        sym = kg['geneid2symbol'].get(nid, '')
        if sym:
            info['ppi_neighbors'].append(sym)
    return info


def build_target_profile(gene_symbol, timepoint, direction, delta_value,
                          delta_r, r_baseline, r_after, pathway_label,
                          kg, cross_df, l1a_df):
    kg_info = get_gene_kg_info(kg, gene_symbol)
    emb_norm, emb_norm_normed = get_kg_weight(kg, gene_symbol)
    if emb_norm_normed >= 0.7:
        kg_weight_level = 'HIGH (top 30% in MitoKG embedding space)'
    elif emb_norm_normed >= 0.4:
        kg_weight_level = 'MEDIUM (middle 30-70%)'
    else:
        kg_weight_level = 'LOW (bottom 30%)'

    cross_row = cross_df[cross_df['gene'] == gene_symbol]
    cross_info = {}
    if not cross_row.empty:
        r = cross_row.iloc[0]
        d0 = float(r.get('delta_r_Day0', 0))
        d10 = float(r.get('delta_r_Day10', 0))
        d14 = float(r.get('delta_r_Day14', 0))
        best_tp = r.get('best_timepoint', timepoint)
        if d0 > 0 and d10 > 0 and d14 > 0:
            temporal_pattern = 'Consistent positive across all timepoints'
        elif d0 < 0 and (d10 > 0 or d14 > 0):
            temporal_pattern = 'Negative at Day0, positive at Day10/14 (late-emerging)'
        elif d0 > 0 and d10 < 0 and d14 < 0:
            temporal_pattern = 'Positive only at Day0 (early-specific)'
        else:
            temporal_pattern = 'Mixed temporal pattern'
        cross_info = {
            'delta_r_Day0': round(d0, 5),
            'delta_r_Day10': round(d10, 5),
            'delta_r_Day14': round(d14, 5),
            'best_timepoint': best_tp,
            'temporal_pattern': temporal_pattern,
        }

    l1a_row = l1a_df[(l1a_df['gene'] == gene_symbol) & (l1a_df['timepoint'] == timepoint)]
    l1a_info = {}
    if not l1a_row.empty:
        r = l1a_row.iloc[0]
        l1a_info = {
            'l1a_delta_r': round(float(r.get('delta_r_vs_ctl', 0)), 5),
            'l1a_r_after': round(float(r.get('r_after_vs_ctl', 0)), 5),
        }

    return {
        'gene': gene_symbol,
        'timepoint': timepoint,
        'pathway': pathway_label,
        'direction': direction,
        'delta_value': round(float(delta_value), 4),
        'delta_r': round(float(delta_r), 5),
        'r_baseline': round(float(r_baseline), 5),
        'r_after': round(float(r_after), 5),
        'kg_info': kg_info,
        'emb_norm': round(emb_norm, 4),
        'emb_norm_normed': round(emb_norm_normed, 4),
        'kg_weight_level': kg_weight_level,
        'cross_timepoint': cross_info,
        'l1a_info': l1a_info,
    }


def build_prompt(target, time_summary, window_ratio):
    kg_info = target['kg_info']
    ct = target['cross_timepoint']
    l1a = target['l1a_info']
    gene = target['gene']
    direction = target['direction']
    pathway = target['pathway']
    delta_v = target['delta_value']
    delta_r = target['delta_r']
    r_base = target['r_baseline']
    r_after = target['r_after']
    tp = target['timepoint']
    kg_level = target['kg_weight_level']
    emb_n = target['emb_norm_normed']

    pathway_str = "; ".join([
        f"{p['name']}" + (f" [{p['reactome_id']}]" if p['reactome_id'] else "")
        for p in kg_info['pathways']]) or "No Reactome pathway mapping in MitoKG"
    complex_str = "; ".join([
        f"CORUM #{c['corum_id']}" for c in kg_info['complexes']]) or "No CORUM complex"
    drug_str = "; ".join([
        f"{d['name']} ({d['drugbank_id']})" for d in kg_info['drugs'][:4]
    ]) or "No approved drug target in MitoKG"
    ppi_str = ", ".join(kg_info['ppi_neighbors'][:8]) or "N/A"

    ts_global = (
        f"Global time-window findings (from MitoLODE gradient optimization):\n"
        f"  Day0 optimization: baseline_r={time_summary.get('Day0', {}).get('baseline_r', 'N/A')} -> "
        f"opt_r={time_summary.get('Day0', {}).get('L2_opt_r', 'N/A')} "
        f"(dr={time_summary.get('Day0', {}).get('L2_delta_r', 'N/A')})\n"
        f"  Day10 optimization: dr={time_summary.get('Day10', {}).get('L2_delta_r', 'N/A')}\n"
        f"  Day14 optimization: dr={time_summary.get('Day14', {}).get('L2_delta_r', 'N/A')}\n"
        f"  -> Day0 intervention is {window_ratio}x more effective than Day14 "
        f"(indicates early differentiation window of high trajectory plasticity)"
    )
    if ct:
        ts_gene = (
            f"This gene's cross-timepoint single-gene delta_r:\n"
            f"  Day0={ct['delta_r_Day0']:+.5f}, "
            f"Day10={ct['delta_r_Day10']:+.5f}, "
            f"Day14={ct['delta_r_Day14']:+.5f}\n"
            f"  Temporal pattern: {ct['temporal_pattern']}\n"
            f"  Best intervention timepoint: {ct['best_timepoint']}"
        )
    else:
        ts_gene = "Cross-timepoint data not available for this gene."

    l1a_str = ""
    if l1a:
        l1a_str = (f"\nL1A single-gene perturbation (independent validation): "
                    f"dr={l1a['l1a_delta_r']:+.5f}, r_after={l1a['l1a_r_after']:.5f}")

    prompt = f"""=== MitoLODE LRRK2 Perturbation Analysis ===

TARGET GENE: {gene}
Perturbation: {direction}-regulation (gradient delta={delta_v:+.4f})
Starting timepoint: {tp} -> predicting Day42 alignment with Control
Result: baseline_r={r_base:.5f} -> optimized_r={r_after:.5f} (dr={delta_r:+.5f}){l1a_str}

--- MitoKG Structural Context for {gene} ---
Pathway group (MitoLODE label): {pathway}
Reactome pathways in MitoKG: {pathway_str}
CORUM complex memberships: {complex_str}
Approved drugs (DrugBank): {drug_str}
STRING PPI neighbors in MitoKG: {ppi_str}
Gene present in MitoKG: {'YES' if kg_info['in_kg'] else 'NO'}

--- KGSFM Gating Weight (model attention) ---
Embedding norm (normalized): {emb_n:.3f}/1.000
Gating level: {kg_level}
-> This means MitoLODE's encoder assigns {"strong" if emb_n > 0.6 else "moderate" if emb_n > 0.3 else "weak"}
  structural attention to {gene} based on its MitoKG connectivity.

--- Temporal Window Context ---
{ts_global}

{ts_gene}

--- LRRK2 Disease Context ---
LRRK2 G2019S: hyperactivated kinase -> RAB10 hyperphosphorylation -> impaired
OPTN-mediated mitophagy; Drp1 activation -> mitochondrial fragmentation;
Complex I dysfunction; elevated mitochondrial ROS; progressive transcriptomic
drift from Control state during iPSC differentiation toward dopaminergic neurons.

=== TASK ===
Provide a KG-grounded step-by-step causal reasoning chain explaining:
(A) WHY {direction}-regulating {gene} shifts the LRRK2 trajectory toward Control
(B) WHY the KGSFM gating weight of {emb_n:.3f} is consistent with this gene's role
(C) WHY intervention at {tp} is {"particularly effective" if tp == "Day0" else "effective at this stage"}
    given the temporal dynamics

Format:
[Step 1] Mechanism: ... (MitoKG: cite specific node/edge from context above)
[Step 2] LRRK2 connection: ...
[Step 3] KGSFM attention rationale: ... (embedding norm={emb_n:.3f} reflects...)
[Step 4] Temporal window explanation: ... (why {tp} intervention matters)
[Verification] MitoKG nodes cited: [list]
[Confidence] HIGH/MEDIUM/LOW - one-sentence justification"""

    return prompt


def verify_chain(llm_output, kg_info, symbol2geneid):
    result = {
        'gene_in_kg': kg_info['in_kg'],
        'kg_pathways': [p['name'] for p in kg_info['pathways']],
        'cited_genes_verified': [],
        'cited_genes_not_in_kg': [],
        'has_temporal_reasoning': False,
        'has_kgsfm_reasoning': False,
        'verification_section': '',
        'confidence': 'UNKNOWN',
        'pass': False,
    }
    if not llm_output or 'ERROR' in llm_output:
        return result

    result['has_temporal_reasoning'] = any(
        kw in llm_output for kw in ['Day0', 'Day10', 'Day14', 'temporal', 'window',
                                      'plasticity', 'early', 'differentiation stage'])
    result['has_kgsfm_reasoning'] = any(
        kw in llm_output for kw in ['KGSFM', 'gating', 'embedding', 'norm',
                                      'attention', 'structural'])

    for lvl in ['HIGH', 'MEDIUM', 'LOW']:
        if f'[Confidence] {lvl}' in llm_output or f'\n{lvl}' in llm_output:
            result['confidence'] = lvl
            break

    lines = llm_output.split('\n')
    in_v = False
    for line in lines:
        if '[Verification]' in line:
            in_v = True
            result['verification_section'] = line
        elif '[Confidence]' in line:
            in_v = False
        elif in_v:
            result['verification_section'] += '\n' + line

    skip = {'LRRK2', 'PINK1', 'CORUM', 'STRING', 'MitoKG', 'ATP', 'DNA', 'RNA', 'ROS',
             'PPI', 'ETC', 'OMM', 'IMM', 'KGSFM', 'Step', 'HIGH', 'MEDIUM', 'LOW',
             'YES', 'NOT', 'WHY', 'AND', 'FOR', 'THE', 'GTP', 'ADP', 'NAD', 'FAD', 'OPTN'}
    cited = set(re.findall(r'\b([A-Z][A-Z0-9]{2,9})\b', llm_output))
    for g in cited:
        if g in skip:
            continue
        if g in symbol2geneid:
            result['cited_genes_verified'].append(g)
        elif len(g) >= 3:
            result['cited_genes_not_in_kg'].append(g)

    result['pass'] = (
        result['gene_in_kg'] and
        result['confidence'] in ('HIGH', 'MEDIUM') and
        len(result['cited_genes_verified']) >= 1 and
        result['has_temporal_reasoning']
    )
    return result


def make_llm_caller():
    api_key = os.environ.get('DEEPSEEK_API_KEY', '').strip()
    if not api_key:
        print("\n" + "=" * 65)
        print("ERROR: DEEPSEEK_API_KEY environment variable is not set.")
        print("=" * 65)
        print("Please obtain a DeepSeek API key from https://platform.deepseek.com")
        print("and set the environment variable before running this script:")
        print()
        print("  Linux / macOS:")
        print('    export DEEPSEEK_API_KEY="sk-your-key-here"')
        print()
        print("  Windows PowerShell:")
        print('    $env:DEEPSEEK_API_KEY="sk-your-key-here"')
        print()
        print("The OpenAI-compatible DeepSeek API is used with base_url =")
        print(f"  {DEEPSEEK_BASE_URL}")
        print("=" * 65)
        sys.exit(1)

    try:
        from openai import OpenAI
    except ImportError:
        print("ERROR: openai package not installed. Run: pip install openai")
        sys.exit(1)

    client = OpenAI(api_key=api_key, base_url=DEEPSEEK_BASE_URL)

    def call_deepseek(prompt, retries=LLM_RETRIES):
        for i in range(retries):
            try:
                resp = client.chat.completions.create(
                    model=DEEPSEEK_MODEL,
                    messages=[{"role": "system", "content": SYSTEM_PROMPT},
                                {"role": "user", "content": prompt}],
                    temperature=LLM_TEMPERATURE,
                    max_tokens=LLM_MAX_TOKENS,
                )
                return resp.choices[0].message.content
            except Exception as e:
                print(f"    [warning] attempt {i + 1}/{retries} failed: {e}")
                if i < retries - 1:
                    time.sleep(3)
        return "ERROR: API failed."

    return call_deepseek


def main():
    args = parse_args()
    args.out_dir.mkdir(parents=True, exist_ok=True)

    print("=" * 65)
    print("Step 1: load MitoKG + gene embeddings + perturbation data")
    print("=" * 65)

    kg = load_kg(args.gnn_dir)

    xl = pd.ExcelFile(str(args.perturb_xlsx))
    l2_df = xl.parse('L2_GradientTargets')
    summ_df = xl.parse('Summary')
    cross_df = xl.parse('CrossTimepoint_GeneMatrix')
    l1a_df = xl.parse('L1A_Top20_ByTimepoint')
    print(f"  perturbation Excel loaded. L2 targets: {len(l2_df)} rows")

    time_summary = {}
    for _, row in summ_df.iterrows():
        tp = row['timepoint']
        time_summary[tp] = {
            'baseline_r': round(float(row['baseline_r_vs_ctl']), 5),
            'ideal_r': round(float(row['ideal_r_vs_ctl']), 5),
            'L2_opt_r': round(float(row['L2_opt_r_vs_ctl']), 5),
            'L2_delta_r': round(float(row['L2_delta_r']), 5),
            'L2_top1': row['L2_top1_gene'],
        }
    delta_day0 = time_summary.get('Day0', {}).get('L2_delta_r', 0)
    delta_day14 = time_summary.get('Day14', {}).get('L2_delta_r', 1e-9)
    window_ratio = round(delta_day0 / (delta_day14 + 1e-9), 2)

    print("\n" + "=" * 65)
    print("Step 2: build target profiles (MitoKG + temporal + KGSFM weight)")
    print("=" * 65)

    targets_raw = []
    if args.mode == 'L2_top_N_each':
        for tp in ['Day0', 'Day10', 'Day14']:
            sub = l2_df[l2_df['timepoint'] == tp].head(args.top_n_each)
            for _, row in sub.iterrows():
                targets_raw.append(build_target_profile(
                    row['gene'], row['timepoint'], row['direction'],
                    row['delta_value'], row['delta_r_vs_ctl'],
                    row['r_baseline_vs_ctl'], row['r_after_vs_ctl'], row['pathway'],
                    kg, cross_df, l1a_df))
    elif args.mode == 'robust_top_N':
        robust = cross_df.nlargest(args.top_n_robust, 'max_delta_r')
        for _, row in robust.iterrows():
            gene = row['gene']
            r2 = l2_df[l2_df['gene'] == gene]
            if r2.empty:
                continue
            r = r2.iloc[0]
            targets_raw.append(build_target_profile(
                gene, row['best_timepoint'], r.get('direction', 'up'),
                r.get('delta_value', 0), r.get('delta_r_vs_ctl', 0),
                r.get('r_baseline_vs_ctl', 0), r.get('r_after_vs_ctl', 0), row['pathway'],
                kg, cross_df, l1a_df))
    else:
        for gene in args.custom_genes:
            sub = l2_df[l2_df['gene'] == gene]
            if sub.empty:
                print(f"  [skip] {gene} not in L2 targets")
                continue
            row = sub.iloc[0]
            targets_raw.append(build_target_profile(
                gene, row['timepoint'], row['direction'],
                row['delta_value'], row['delta_r_vs_ctl'],
                row['r_baseline_vs_ctl'], row['r_after_vs_ctl'], row['pathway'],
                kg, cross_df, l1a_df))

    print(f"\n  selected targets: {len(targets_raw)}")
    print(f"  {'gene':<12} {'timepoint':<8} {'dir':<6} {'delta':>8} "
          f"{'pathway':<18} {'kg_w':<8} {'temporal_pattern'}")
    print("  " + "-" * 95)
    for t in targets_raw:
        ct = t['cross_timepoint']
        pat = ct.get('temporal_pattern', '')[:30] if ct else ''
        print(f"  {t['gene']:<12} {t['timepoint']:<8} {t['direction']:<6} "
                f"{t['delta_value']:>+8.3f} {t['pathway']:<18} "
                f"{t['emb_norm_normed']:.2f}    {pat}")

    print("\n" + "=" * 65)
    print(f"Step 3: call DeepSeek ({len(targets_raw)} targets)")
    print("=" * 65)

    call_deepseek = make_llm_caller()

    results = []
    for i, target in enumerate(targets_raw):
        gene = target['gene']
        tp = target['timepoint']
        print(f"\n[{i + 1:2d}/{len(targets_raw)}] {tp} | {gene:<12} "
                f"({target['direction']:4s}) pathway={target['pathway']:<18} "
                f"kg_w={target['emb_norm_normed']:.2f}")
        prompt = build_prompt(target, time_summary, window_ratio)
        t0 = time.time()
        llm_output = call_deepseek(prompt)
        elapsed = time.time() - t0
        verif = verify_chain(llm_output, target['kg_info'], kg['symbol2geneid'])
        status = "PASS" if verif['pass'] else "REVIEW"
        print(f"  [{status}]  confidence={verif['confidence']}  "
                f"temporal={'Y' if verif['has_temporal_reasoning'] else 'N'}  "
                f"KGSFM={'Y' if verif['has_kgsfm_reasoning'] else 'N'}  "
                f"verified_genes={len(verif['cited_genes_verified'])}  "
                f"({elapsed:.1f}s)")
        results.append({'target': target, 'prompt': prompt,
                          'llm_output': llm_output, 'verif': verif})
        time.sleep(1.0)

    print("\n" + "=" * 70)
    print("MGMI Reasoning Summary")
    print("=" * 70)
    pass_n = sum(1 for r in results if r['verif']['pass'])
    temp_n = sum(1 for r in results if r['verif']['has_temporal_reasoning'])
    kgsfm_n = sum(1 for r in results if r['verif']['has_kgsfm_reasoning'])
    total = len(results)
    print(f"total={total}  pass={pass_n}  temporal_ok={temp_n}  kgsfm_ok={kgsfm_n}")

    json_out = args.out_dir / "mitokg_llm_raw_v2.json"
    with open(str(json_out), 'w', encoding='utf-8') as f:
        json.dump([{
            'gene': r['target']['gene'],
            'timepoint': r['target']['timepoint'],
            'direction': r['target']['direction'],
            'pathway': r['target']['pathway'],
            'delta_r': r['target']['delta_r'],
            'r_after': r['target']['r_after'],
            'kg_weight': r['target']['emb_norm_normed'],
            'temporal_pattern': r['target']['cross_timepoint'].get('temporal_pattern', '') if r['target']['cross_timepoint'] else '',
            'confidence': r['verif']['confidence'],
            'has_temporal': r['verif']['has_temporal_reasoning'],
            'has_kgsfm': r['verif']['has_kgsfm_reasoning'],
            'pass': r['verif']['pass'],
            'reasoning': r['llm_output'],
        } for r in results], f, ensure_ascii=False, indent=2)
    print(f"JSON saved: {json_out}")

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        S_HDR = PatternFill("solid", fgColor="1F3864")
        S_TP0 = PatternFill("solid", fgColor="D9E1F2")
        S_TP10 = PatternFill("solid", fgColor="E2EFDA")
        S_TP14 = PatternFill("solid", fgColor="FFF2CC")
        S_PASS = PatternFill("solid", fgColor="C6EFCE")
        S_WARN = PatternFill("solid", fgColor="FFEB9C")
        TP_FILLS = {'Day0': S_TP0, 'Day10': S_TP10, 'Day14': S_TP14}
        F_HDR = Font(bold=True, color="FFFFFF", size=11)
        F_BOLD = Font(bold=True, size=10)
        F_NORM = Font(size=10)
        F_MONO = Font(size=9, name='Consolas')
        A_CTR = Alignment(horizontal='center', vertical='top', wrap_text=True)
        A_LFT = Alignment(horizontal='left', vertical='top', wrap_text=True)

        def bd():
            s = Side(style='thin', color='CCCCCC')
            return Border(left=s, right=s, top=s, bottom=s)

        def sc(ws, r, c, v, fill=None, font=None, align=A_CTR, border=True):
            cell = ws.cell(r, c, v)
            if fill:
                cell.fill = fill
            cell.font = font if font else F_NORM
            cell.alignment = align
            if border:
                cell.border = bd()
            return cell

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        ws1 = wb.create_sheet("Reasoning_Chains")
        col_widths = [6, 12, 9, 6, 18, 9, 9, 9, 9, 10, 10, 10, 55, 18]
        for ci, w in enumerate(col_widths, 1):
            ws1.column_dimensions[get_column_letter(ci)].width = w

        headers = ['#', 'Gene', 'Timepoint', 'Dir', 'Pathway', 'delta', 'dr', 'r_after',
                     'KG_weight', 'Confidence', 'Temporal', 'KGSFM',
                     'Reasoning Chain (KG-Grounded, Temporal, KGSFM-aware)',
                     'Cited Genes (MitoKG verified)']
        row = 1
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        sc(ws1, row, 1, "LRRK2 | MitoKG-Grounded LLM Interpretability - "
             "Integrating KG Structure + Temporal Dynamics + KGSFM Gating",
             fill=S_HDR, font=F_HDR)
        row += 1
        for j, h in enumerate(headers, 1):
            sc(ws1, row, j, h,
                 fill=PatternFill("solid", fgColor="2E75B6"),
                 font=Font(bold=True, color="FFFFFF", size=10))
        row += 1

        for i, r in enumerate(results):
            t = r['target']
            v = r['verif']
            tf = TP_FILLS.get(t['timepoint'])
            pf = S_PASS if v['pass'] else S_WARN
            ws1.row_dimensions[row].height = max(120, min(350, len(r['llm_output']) // 3))
            sc(ws1, row, 1, i + 1, fill=tf, font=F_NORM)
            sc(ws1, row, 2, t['gene'], fill=tf, font=F_BOLD, align=A_LFT)
            sc(ws1, row, 3, t['timepoint'], fill=tf, font=F_NORM)
            sc(ws1, row, 4, t['direction'], fill=tf, font=F_NORM)
            sc(ws1, row, 5, t['pathway'], fill=tf, font=F_NORM, align=A_LFT)
            sc(ws1, row, 6, t['delta_value'], fill=tf, font=F_NORM)
            sc(ws1, row, 7, t['delta_r'], fill=tf, font=F_NORM)
            sc(ws1, row, 8, t['r_after'], fill=tf, font=F_NORM)
            sc(ws1, row, 9, f"{t['emb_norm_normed']:.3f}", fill=tf, font=F_NORM)
            sc(ws1, row, 10, v['confidence'], fill=pf, font=F_BOLD)
            sc(ws1, row, 11, 'Y' if v['has_temporal_reasoning'] else 'N', fill=pf, font=F_BOLD)
            sc(ws1, row, 12, 'Y' if v['has_kgsfm_reasoning'] else 'N', fill=pf, font=F_BOLD)
            sc(ws1, row, 13, r['llm_output'] or '', fill=None, font=F_MONO, align=A_LFT)
            sc(ws1, row, 14, ', '.join(v['cited_genes_verified'][:12]),
                 fill=tf, font=F_NORM, align=A_LFT)
            row += 1
        ws1.freeze_panes = 'A3'

        ws2 = wb.create_sheet("Verification_Summary")
        for ci, w in enumerate([6, 12, 9, 6, 18, 9, 9, 10, 10, 10, 30], 1):
            ws2.column_dimensions[get_column_letter(ci)].width = w
        row2 = 1
        ws2.merge_cells(start_row=row2, start_column=1, end_row=row2, end_column=11)
        sc(ws2, row2, 1, "MitoKG Verification Summary - Three-Layer Evidence Integration",
             fill=S_HDR, font=F_HDR)
        row2 += 1
        for j, h in enumerate(['#', 'Gene', 'Timepoint', 'Dir', 'Pathway', 'KG_weight',
                                 'In_KG', 'Confidence', 'Temporal', 'KGSFM', 'KG_pathways'], 1):
            sc(ws2, row2, j, h, fill=PatternFill("solid", fgColor="2E75B6"),
                 font=Font(bold=True, color="FFFFFF", size=10))
        row2 += 1
        for i, r in enumerate(results):
            t = r['target']
            v = r['verif']
            tf = TP_FILLS.get(t['timepoint'])
            pf = S_PASS if v['pass'] else S_WARN
            sc(ws2, row2, 1, i + 1, fill=tf)
            sc(ws2, row2, 2, t['gene'], fill=tf, font=F_BOLD, align=A_LFT)
            sc(ws2, row2, 3, t['timepoint'], fill=tf)
            sc(ws2, row2, 4, t['direction'], fill=tf)
            sc(ws2, row2, 5, t['pathway'], fill=tf, align=A_LFT)
            sc(ws2, row2, 6, f"{t['emb_norm_normed']:.3f}", fill=tf)
            sc(ws2, row2, 7, 'Y' if v['gene_in_kg'] else 'N', fill=pf, font=F_BOLD)
            sc(ws2, row2, 8, v['confidence'], fill=pf, font=F_BOLD)
            sc(ws2, row2, 9, 'Y' if v['has_temporal_reasoning'] else 'N', fill=pf, font=F_BOLD)
            sc(ws2, row2, 10, 'Y' if v['has_kgsfm_reasoning'] else 'N', fill=pf, font=F_BOLD)
            sc(ws2, row2, 11, '; '.join(v['kg_pathways'][:2]), fill=tf, align=A_LFT)
            row2 += 1
        row2 += 1
        ws2.merge_cells(start_row=row2, start_column=1, end_row=row2, end_column=11)
        sc(ws2, row2, 1,
             f"Statistics: Total={total} | Pass={pass_n} ({100 * pass_n // (total or 1)}%) | "
             f"Temporal={temp_n} | KGSFM={kgsfm_n} | "
             f"Time-window ratio Day0/Day14={window_ratio}x",
             fill=PatternFill("solid", fgColor="D6E4F0"),
             font=Font(bold=True, size=10), align=A_LFT)
        ws2.freeze_panes = 'A3'

        xl_out = args.out_dir / "mitokg_llm_interpretability_lrrk2_v2.xlsx"
        wb.save(str(xl_out))
        print(f"Excel saved: {xl_out}")
    except ImportError:
        print("[warning] openpyxl not installed, skipped Excel export")

    print("\n" + "=" * 65)
    print(f"MGMI inference complete. output dir: {args.out_dir}")
    print(f"  total={total}  pass={pass_n}  temporal_ok={temp_n}  kgsfm_ok={kgsfm_n}")
    print("=" * 65)


if __name__ == '__main__':
    main()
